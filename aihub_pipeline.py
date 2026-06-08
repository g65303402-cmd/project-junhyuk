"""
AI Hub 데이터 → 박준혁 스타일 변환 파이프라인 v1
-------------------------------------------------
감성대화(018) xlsx + 공감형대화(046) JSON
→ Claude API로 박준혁 반말 변환
→ 기존 train.json에 병합

서버에서 실행 예시:
  python3 aihub_pipeline.py \
    --감성대화_dir  /path/to/018_감성대화  \
    --공감형대화_dir /path/to/046_공감형_대화 \
    --existing_train train.json \
    --output merged_train.json \
    --target 400
"""

import os, sys, json, zipfile, io, random, time, argparse
from collections import defaultdict, Counter

try:
    import openpyxl
except ImportError:
    os.system("pip install openpyxl --break-system-packages -q")
    import openpyxl

# ──────────────────────────────────────────────
# 필터 설정
# ──────────────────────────────────────────────
EMOTION_FILTER = {"불안", "분노", "상처", "슬픔", "당황"}
SITUATION_FILTER = {
    "대인관계", "진로,취업,직장", "가족관계", "연애,결혼,출산",
    "대인관계(부부, 자녀)", "학업 및 진로", "재정", "일상",
}
EMPATHY_EMOTION_FILTER = {"슬픔", "불안", "분노", "상처", "당황"}

CATEGORY_MAP = {
    "진로,취업,직장": "직장 및 학업 스트레스",
    "학업 및 진로":   "직장 및 학업 스트레스",
    "대인관계":       "인간관계 갈등",
    "대인관계(부부, 자녀)": "인간관계 갈등",
    "가족관계":       "인간관계 갈등",
    "연애,결혼,출산": "연애 및 이별 후유증",
    "재정":           "직장 및 학업 스트레스",
    "분노":           "자존감 및 자기비판",
    "상처":           "자존감 및 자기비판",
    "슬픔":           "무기력 및 공허함",
    "불안":           "무기력 및 공허함",
    "당황":           "인간관계 갈등",
}

# ──────────────────────────────────────────────
# Claude API 프롬프트
# ──────────────────────────────────────────────
SYSTEM_PROMPT = """너는 박준혁이야. 29세, 전직 군상담관 출신 자원봉사 상담사.
다음 규칙을 반드시 지켜.

말투:
- 반말만 사용. "~요", "~습니다", "~세요" 절대 금지
- 짧고 간결하게. 2~3문장 이내
- 딱딱한 상담 용어 금지

응답 구조 (3단계 반드시 순서대로):
1. 인정 — 상황을 있는 그대로 받아들이는 한 문장
   예: "그렇구나", "그럴 수밖에 없었겠다", "충분히 그럴 수 있어", "그게 쉽지 않았겠다", "많이 힘들었겠다", "그 마음 이해해", "그랬던 거야"
   — 매번 다르게, 같은 표현 연속 금지
2. 공감 — 감정의 무게를 감각적으로 짚어주는 1문장 (단어 나열 말고 느낌으로)
3. 열린 질문 — 딱 하나만

절대 금지: 존댓말 / "힘내" 단독 / 진단 / 조언 / 질문 2개 이상"""

USER_TEMPLATE = "사용자: {user_text}\n\n박준혁 응답 (반말, 2~3문장):"


# ──────────────────────────────────────────────
# 1. 감성대화 xlsx 추출
# ──────────────────────────────────────────────
def find_zip(base_dir, keyword_in_path):
    """base_dir 아래에서 경로에 keyword가 포함된 zip 찾기"""
    for root, dirs, files in os.walk(base_dir):
        for f in files:
            if f.endswith('.zip') and keyword_in_path in root:
                return os.path.join(root, f)
    return None


def extract_from_xlsx(감성대화_dir, max_per_emotion=80):
    print("📂 감성대화(xlsx) 로딩 중...")

    # Training 원천데이터 zip 찾기
    zip_path = find_zip(감성대화_dir, 'Training')
    if not zip_path:
        print("  ⚠️ Training zip 못 찾음, 건너뜀")
        return []

    print(f"  zip: {os.path.basename(zip_path)}")

    with zipfile.ZipFile(zip_path) as zf:
        xlsx_names = [n for n in zf.namelist() if n.endswith('.xlsx')]
        if not xlsx_names:
            print("  ⚠️ xlsx 없음")
            return []
        with zf.open(xlsx_names[0]) as f:
            wb = openpyxl.load_workbook(io.BytesIO(f.read()))
            ws = wb.active
            print(f"  총 {ws.max_row:,}행")

            samples = defaultdict(list)
            for r in range(2, ws.max_row + 1):
                situation = ws.cell(r, 4).value or ""
                emotion   = ws.cell(r, 6).value or ""
                person1   = ws.cell(r, 8).value
                if not person1:
                    continue
                if emotion not in EMOTION_FILTER:
                    continue
                if situation not in SITUATION_FILTER:
                    continue
                category = CATEGORY_MAP.get(situation) or CATEGORY_MAP.get(emotion, "일상 및 기타")
                samples[emotion].append({
                    "user": str(person1).strip(),
                    "category": category,
                })

    result = []
    for emotion, items in samples.items():
        random.shuffle(items)
        result.extend(items[:max_per_emotion])

    print(f"  필터 후 추출: {len(result)}개")
    return result


# ──────────────────────────────────────────────
# 2. 공감형대화 JSON 추출
# ──────────────────────────────────────────────
def extract_from_empathy(공감형대화_dir, max_per_zip=12):
    print("📂 공감형대화(JSON) 로딩 중...")

    tl_zips = []
    for root, dirs, files in os.walk(공감형대화_dir):
        for f in files:
            if f.startswith('TL_') and f.endswith('.zip'):
                tl_zips.append(os.path.join(root, f))

    print(f"  TL zip 수: {len(tl_zips)}")
    result = []

    for zip_path in tl_zips:
        try:
            with zipfile.ZipFile(zip_path) as zf:
                json_files = [n for n in zf.namelist() if n.endswith('.json')]
                random.shuffle(json_files)
                count = 0
                for jf in json_files:
                    if count >= max_per_zip:
                        break
                    try:
                        with zf.open(jf) as f:
                            data = json.load(f)
                        emotion = data.get("info", {}).get("speaker_emotion", "")
                        if emotion not in EMPATHY_EMOTION_FILTER:
                            continue
                        for utt in data.get("utterances", []):
                            if utt.get("role") == "speaker":
                                text = utt.get("text", "").strip()
                                if len(text) >= 10:
                                    result.append({
                                        "user": text,
                                        "category": CATEGORY_MAP.get(emotion, "일상 및 기타"),
                                    })
                                    count += 1
                                    break
                    except Exception:
                        continue
        except Exception as e:
            print(f"  ⚠️ {os.path.basename(zip_path)}: {e}")

    print(f"  추출 완료: {len(result)}개")
    return result


# ──────────────────────────────────────────────
# 3. Claude API 변환
# ──────────────────────────────────────────────
def call_claude(user_text):
    import urllib.request, urllib.error
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        raise RuntimeError("ANTHROPIC_API_KEY 환경변수가 설정되지 않았습니다.\n  export ANTHROPIC_API_KEY='sk-ant-...'")

    payload = json.dumps({
        "model": "claude-opus-4-5",
        "max_tokens": 1000,
        "system": SYSTEM_PROMPT,
        "messages": [{"role": "user", "content": USER_TEMPLATE.format(user_text=user_text)}]
    }).encode("utf-8")

    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages",
        data=payload,
        headers={
            "Content-Type": "application/json",
            "anthropic-version": "2023-06-01",
            "x-api-key": api_key,
        },
        method="POST"
    )
    with urllib.request.urlopen(req, timeout=30) as resp:
        data = json.loads(resp.read())
        return data["content"][0]["text"].strip()


def convert_batch(samples, target_count, delay=0.5, save_path="converted_temp.json"):
    random.shuffle(samples)
    samples = samples[:target_count]

    print(f"\n🤖 Claude API 변환 시작 ({len(samples)}개)...")
    results = []
    failed = 0

    for i, item in enumerate(samples):
        try:
            assistant_text = call_claude(item["user"])
            results.append({
                "category": item["category"],
                "user": item["user"],
                "assistant": assistant_text,
            })
            if (i + 1) % 10 == 0:
                print(f"  {i+1}/{len(samples)} 완료 ({len(results)} 성공)")
                # 중간 저장
                with open(save_path, "w", encoding="utf-8") as f:
                    json.dump(results, f, ensure_ascii=False, indent=2)
        except Exception as e:
            failed += 1
            if failed <= 5:
                print(f"  ⚠️ 실패 ({i+1}): {e}")
        time.sleep(delay)

    print(f"  변환 완료: {len(results)}성공 / {failed}실패")
    # 최종 저장
    with open(save_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    return results


# ──────────────────────────────────────────────
# 4. 병합
# ──────────────────────────────────────────────
def merge(existing_path, new_samples, output_path):
    with open(existing_path, encoding="utf-8") as f:
        existing = json.load(f)

    existing_users = {item["user"].strip() for item in existing}
    deduped, dupes = [], 0
    for item in new_samples:
        if item["user"].strip() not in existing_users:
            deduped.append(item)
            existing_users.add(item["user"].strip())
        else:
            dupes += 1

    merged = existing + deduped
    random.shuffle(merged)

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(merged, f, ensure_ascii=False, indent=2)

    print(f"\n✅ 병합 완료!")
    print(f"  기존 {len(existing)} + 신규 {len(deduped)} (중복 {dupes}개 제거) = {len(merged)}개")
    print(f"  저장: {output_path}")

    cat = Counter(item["category"] for item in merged)
    print("\n📊 카테고리 분포:")
    for c, n in sorted(cat.items(), key=lambda x: -x[1]):
        print(f"  {c}: {n}개")
    return merged


# ──────────────────────────────────────────────
# 메인
# ──────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--감성대화_dir",  required=True)
    parser.add_argument("--공감형대화_dir", required=True)
    parser.add_argument("--existing_train", required=True)
    parser.add_argument("--output", default="merged_train.json")
    parser.add_argument("--target", type=int, default=400)
    parser.add_argument("--dry_run", action="store_true", help="API 호출 없이 추출만 확인")
    parser.add_argument("--delay", type=float, default=0.3, help="API 호출 간격(초)")
    args = parser.parse_args()

    random.seed(42)

    xlsx_target    = int(args.target * 0.6)
    empathy_target = args.target - xlsx_target

    xlsx_samples    = extract_from_xlsx(args.감성대화_dir, max_per_emotion=xlsx_target // 5 + 20)
    empathy_samples = extract_from_empathy(args.공감형대화_dir, max_per_zip=12)

    all_samples = xlsx_samples + empathy_samples
    print(f"\n후보 총 {len(all_samples)}개 (xlsx {len(xlsx_samples)} + 공감 {len(empathy_samples)})")

    if args.dry_run:
        print("\n[dry_run] 샘플 5개:")
        for s in random.sample(all_samples, min(5, len(all_samples))):
            print(f"  [{s['category']}] {s['user'][:70]}")
        print(f"\nAPI 변환 예정: {min(args.target, len(all_samples))}개")
        return

    converted = convert_batch(all_samples, args.target, delay=args.delay,
                               save_path="converted_temp.json")
    merge(args.existing_train, converted, args.output)


if __name__ == "__main__":
    main()